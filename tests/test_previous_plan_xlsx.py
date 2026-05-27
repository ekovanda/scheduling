"""Tests for parsing previous quarter xlsx into carry-forward context."""

import io
from datetime import date

import openpyxl
import pytest

from app.scheduler.file_loader import FileFormatError, parse_previous_plan_xlsx
from app.scheduler.models import (
    Beruf,
    PreviousPlanContext,
    ShiftType,
    Staff,
    build_previous_context_from_xlsx,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_staff(
    identifier: str,
    beruf: Beruf = Beruf.TFA,
    hours: int = 40,
    name: str | None = None,
) -> Staff:
    return Staff(
        name=name or identifier,
        identifier=identifier,
        adult=True,
        hours=hours,
        beruf=beruf,
        reception=True,
        nd_possible=True,
        nd_alone=False,
        nd_max_consecutive=3,
        nd_exceptions=[],
    )


def _build_xlsx(rows: list[dict]) -> io.BytesIO:
    """Build an in-memory xlsx with the long-format previous-plan schema."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Datum", "Wochentag", "Schicht", "Mitarbeiter", "Paarweise"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["Datum"],
            row.get("Wochentag", ""),
            row["Schicht"],
            row["Mitarbeiter"],
            row["Paarweise"],
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_two_sheet_xlsx(
    night_rows: list[dict],
    weekend_rows: list[dict],
) -> io.BytesIO:
    """Build an in-memory two-sheet xlsx matching the current export format."""
    headers = ["Datum", "Wochentag", "Schicht", "Mitarbeiter", "Paarweise"]
    wb = openpyxl.Workbook()

    ws_nights = wb.active
    ws_nights.title = "Nachtdienste"
    ws_nights.append(headers)
    for row in night_rows:
        ws_nights.append([row["Datum"], row.get("Wochentag", ""), row["Schicht"], row["Mitarbeiter"], row["Paarweise"]])

    ws_weekends = wb.create_sheet(title="Wochenenddienste")
    ws_weekends.append(headers)
    for row in weekend_rows:
        ws_weekends.append([row["Datum"], row.get("Wochentag", ""), row["Schicht"], row["Mitarbeiter"], row["Paarweise"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Tests: parse_previous_plan_xlsx
# ---------------------------------------------------------------------------

class TestParsePreviousPlanXlsx:
    def test_solo_night_shift(self) -> None:
        buf = _build_xlsx([
            {"Datum": "01.04.2026", "Schicht": "N_Mi-Do", "Mitarbeiter": "SG", "Paarweise": "Nein"},
        ])
        rows = parse_previous_plan_xlsx(buf)
        assert len(rows) == 1
        assert rows[0]["shift_date"] == date(2026, 4, 1)
        assert rows[0]["shift_type"] == "N_Mi-Do"
        assert rows[0]["staff_identifier"] == "SG"
        assert rows[0]["is_paired"] is False

    def test_paired_night_shift_two_rows(self) -> None:
        """Paired shifts come as two separate rows (one per person)."""
        buf = _build_xlsx([
            {"Datum": "01.04.2026", "Schicht": "N_Mi-Do", "Mitarbeiter": "SG", "Paarweise": "Ja"},
            {"Datum": "01.04.2026", "Schicht": "N_Mi-Do", "Mitarbeiter": "JK", "Paarweise": "Ja"},
        ])
        rows = parse_previous_plan_xlsx(buf)
        assert len(rows) == 2
        assert all(r["is_paired"] is True for r in rows)
        identifiers = {r["staff_identifier"] for r in rows}
        assert identifiers == {"SG", "JK"}

    def test_weekend_shifts(self) -> None:
        buf = _build_xlsx([
            {"Datum": "04.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "AA", "Paarweise": "Nein"},
            {"Datum": "05.04.2026", "Schicht": "So_8-20", "Mitarbeiter": "BB", "Paarweise": "Nein"},
        ])
        rows = parse_previous_plan_xlsx(buf)
        assert len(rows) == 2
        assert rows[0]["shift_type"] == "Sa_10-22"
        assert rows[0]["shift_date"] == date(2026, 4, 4)
        assert rows[1]["shift_type"] == "So_8-20"

    def test_weekend_violation_all_rows_preserved(self) -> None:
        """More than the usual 3 weekend slots on one day must all be parsed."""
        buf = _build_xlsx([
            {"Datum": "04.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "AA", "Paarweise": "Nein"},
            {"Datum": "04.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "BB", "Paarweise": "Nein"},
            {"Datum": "04.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "CC", "Paarweise": "Nein"},
            {"Datum": "04.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "DD", "Paarweise": "Nein"},
        ])
        rows = parse_previous_plan_xlsx(buf)
        assert len(rows) == 4
        assert {r["staff_identifier"] for r in rows} == {"AA", "BB", "CC", "DD"}

    def test_blank_rows_skipped(self) -> None:
        buf = _build_xlsx([
            {"Datum": "01.04.2026", "Schicht": "N_Mi-Do", "Mitarbeiter": "SG", "Paarweise": "Nein"},
            {"Datum": "", "Schicht": "", "Mitarbeiter": "", "Paarweise": ""},
        ])
        rows = parse_previous_plan_xlsx(buf)
        assert len(rows) == 1

    def test_iso_date_format_accepted(self) -> None:
        buf = _build_xlsx([
            {"Datum": "2026-04-01", "Schicht": "N_Mi-Do", "Mitarbeiter": "SG", "Paarweise": "Nein"},
        ])
        rows = parse_previous_plan_xlsx(buf)
        assert rows[0]["shift_date"] == date(2026, 4, 1)

    def test_missing_required_column_raises(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Datum", "Schicht", "Paarweise"])  # missing Mitarbeiter
        ws.append(["01.04.2026", "N_Mi-Do", "Nein"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        with pytest.raises(FileFormatError):
            parse_previous_plan_xlsx(buf)

    def test_two_sheet_xlsx_parses_all_assignments(self) -> None:
        """Two-sheet export format: nights sheet + weekends sheet → all rows recovered."""
        night_rows = [
            {"Datum": "01.04.2026", "Schicht": "N_Mi-Do", "Mitarbeiter": "SG", "Paarweise": "Nein"},
            {"Datum": "07.04.2026", "Schicht": "N_Di-Mi", "Mitarbeiter": "JK", "Paarweise": "Ja"},
        ]
        weekend_rows = [
            {"Datum": "04.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "AA", "Paarweise": "Nein"},
            {"Datum": "05.04.2026", "Schicht": "So_8-20", "Mitarbeiter": "BB", "Paarweise": "Nein"},
        ]
        buf = _build_two_sheet_xlsx(night_rows, weekend_rows)
        rows = parse_previous_plan_xlsx(buf)

        assert len(rows) == 4
        identifiers = {r["staff_identifier"] for r in rows}
        assert identifiers == {"SG", "JK", "AA", "BB"}
        night_parsed = [r for r in rows if r["shift_type"].startswith("N_")]
        weekend_parsed = [r for r in rows if not r["shift_type"].startswith("N_")]
        assert len(night_parsed) == 2
        assert len(weekend_parsed) == 2


# ---------------------------------------------------------------------------
# Tests: build_previous_context_from_xlsx
# ---------------------------------------------------------------------------

class TestBuildPreviousContextFromXlsx:
    def test_returns_previous_plan_context(self) -> None:
        buf = _build_xlsx([
            {"Datum": "01.04.2026", "Schicht": "N_Mi-Do", "Mitarbeiter": "SG", "Paarweise": "Nein"},
            {"Datum": "05.04.2026", "Schicht": "So_8-20", "Mitarbeiter": "SG", "Paarweise": "Nein"},
        ])
        staff = [_make_staff("SG")]
        ctx = build_previous_context_from_xlsx(buf, staff)
        assert isinstance(ctx, PreviousPlanContext)

    def test_quarter_dates_auto_detected(self) -> None:
        buf = _build_xlsx([
            {"Datum": "01.04.2026", "Schicht": "N_Mi-Do", "Mitarbeiter": "SG", "Paarweise": "Nein"},
            {"Datum": "30.06.2026", "Schicht": "N_Do-Fr", "Mitarbeiter": "SG", "Paarweise": "Nein"},
        ])
        staff = [_make_staff("SG")]
        ctx = build_previous_context_from_xlsx(buf, staff)
        assert ctx.quarter_start == date(2026, 4, 1)
        assert ctx.quarter_end == date(2026, 6, 30)

    def test_trailing_assignments_cover_last_21_days(self) -> None:
        # Explicitly place assignments across the quarter; only the last 21 days
        # of the max date should appear in trailing_assignments.
        rows = [
            # Early in quarter – should NOT be in trailing
            {"Datum": "01.04.2026", "Schicht": "N_Mi-Do", "Mitarbeiter": "SG", "Paarweise": "Nein"},
            {"Datum": "01.05.2026", "Schicht": "N_Fr-Sa", "Mitarbeiter": "SG", "Paarweise": "Nein"},
            # Within last 21 days of June 30 (cutoff = June 10) – should BE in trailing
            {"Datum": "15.06.2026", "Schicht": "N_So-Mo", "Mitarbeiter": "SG", "Paarweise": "Nein"},
            {"Datum": "30.06.2026", "Schicht": "N_Di-Mi", "Mitarbeiter": "SG", "Paarweise": "Nein"},
        ]
        buf = _build_xlsx(rows)
        staff = [_make_staff("SG")]
        ctx = build_previous_context_from_xlsx(buf, staff)
        # quarter_end = 30.06.2026, cutoff = 30.06 - 20 days = 10.06.2026
        cutoff = date(2026, 6, 10)
        assert ctx.quarter_end == date(2026, 6, 30)
        trailing_dates = {ta.shift_date for ta in ctx.trailing_assignments}
        assert date(2026, 6, 15) in trailing_dates
        assert date(2026, 6, 30) in trailing_dates
        assert date(2026, 4, 1) not in trailing_dates
        assert date(2026, 5, 1) not in trailing_dates
        for ta in ctx.trailing_assignments:
            assert ta.shift_date >= cutoff

    def test_carry_forward_delta_computed(self) -> None:
        """Two TFA with different workloads get opposite-sign deltas."""
        buf = _build_xlsx([
            # A: 2 weekend shifts
            {"Datum": "04.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "A", "Paarweise": "Nein"},
            {"Datum": "11.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "A", "Paarweise": "Nein"},
            # B: 0 weekend shifts (one night only, same date)
            {"Datum": "01.04.2026", "Schicht": "N_Mi-Do", "Mitarbeiter": "B", "Paarweise": "Nein"},
        ])
        staff = [_make_staff("A"), _make_staff("B")]
        ctx = build_previous_context_from_xlsx(buf, staff)
        by_id = {e.identifier: e for e in ctx.carry_forward}
        assert by_id["A"].carry_forward_delta > by_id["B"].carry_forward_delta

    def test_weekend_violation_counted_in_delta(self) -> None:
        """4 weekend entries for one person on one day → all counted in delta."""
        buf = _build_xlsx([
            {"Datum": "04.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "A", "Paarweise": "Nein"},
            {"Datum": "11.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "A", "Paarweise": "Nein"},
            {"Datum": "18.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "A", "Paarweise": "Nein"},
            {"Datum": "25.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "A", "Paarweise": "Nein"},
            # B has 1 weekend shift (baseline)
            {"Datum": "04.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "B", "Paarweise": "Nein"},
        ])
        staff = [_make_staff("A"), _make_staff("B")]
        ctx = build_previous_context_from_xlsx(buf, staff)
        by_id = {e.identifier: e for e in ctx.carry_forward}
        assert by_id["A"].weekend_shifts == 4
        assert by_id["B"].weekend_shifts == 1
        assert by_id["A"].carry_forward_delta > by_id["B"].carry_forward_delta

    def test_unknown_staff_in_xlsx_ignored_in_carry_forward(self) -> None:
        """Staff in xlsx but not in the new quarter's staff list produce no entry."""
        buf = _build_xlsx([
            {"Datum": "01.04.2026", "Schicht": "N_Mi-Do", "Mitarbeiter": "GHOST", "Paarweise": "Nein"},
            {"Datum": "01.04.2026", "Schicht": "Sa_10-22", "Mitarbeiter": "SG", "Paarweise": "Nein"},
        ])
        staff = [_make_staff("SG")]
        ctx = build_previous_context_from_xlsx(buf, staff)
        identifiers = {e.identifier for e in ctx.carry_forward}
        assert "GHOST" not in identifiers
        assert "SG" in identifiers

    def test_empty_xlsx_raises(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Datum", "Wochentag", "Schicht", "Mitarbeiter", "Paarweise"])
        # No data rows
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        with pytest.raises(ValueError, match="no valid assignment rows"):
            build_previous_context_from_xlsx(buf, [_make_staff("SG")])
